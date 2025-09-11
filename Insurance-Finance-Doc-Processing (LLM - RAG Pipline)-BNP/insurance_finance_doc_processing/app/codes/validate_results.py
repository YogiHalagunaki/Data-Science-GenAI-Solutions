import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def validate_result(result):

    # Load actual JSON from file
    with open("/home/yogi/Desktop/Python_Code/assignment-bnp/Genel_Energy_Trades.json", "r") as f:
        actual_data = json.load(f)

    # Handle result: string or dict
    if isinstance(result, str):
        result_data = json.loads(result)
    else:
        result_data = result

    # Flatten trades
    actual_trades = {trade["TradeID"]: trade for trade in actual_data["trades"]}
    result_trades = {trade["TradeID"]: trade for trade in result_data["trades"]}

    rows = []

    # Compare values
    for trade_id, actual_trade in actual_trades.items():
        result_trade = result_trades.get(trade_id, {})

        for key, actual_value in actual_trade.items():
            result_value = result_trade.get(key, "MISSING")

            status = "MATCH" if actual_value == result_value else "NO MATCH"

            rows.append({
                "TradeID": trade_id,
                "label": key,
                "actual_value": actual_value,
                "result_value": result_value,
                "status": status
            })

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Save Excel to specific path
    excel_file = "/home/yogi/Desktop/Python_Code/assignment-bnp/trade_validation.xlsx"
    df.to_excel(excel_file, index=False)

    # Apply conditional formatting
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(excel_file)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red

    status_col = df.columns.get_loc("status") + 1
    actual_col = df.columns.get_loc("actual_value") + 1
    result_col = df.columns.get_loc("result_value") + 1

    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=status_col)
        actual_cell = ws.cell(row=row, column=actual_col)
        result_cell = ws.cell(row=row, column=result_col)

        if status_cell.value == "MATCH":
            status_cell.fill = green_fill
            actual_cell.fill = green_fill
            result_cell.fill = green_fill
        else:
            status_cell.fill = red_fill
            actual_cell.fill = red_fill
            result_cell.fill = red_fill

    wb.save(excel_file)

    print(f"Validation XL written to {excel_file}")

